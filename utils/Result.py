import os
import sys
sys.path.append('C:/Users/ekans/OneDrive/Desktop/OUR/ris')
import environment as environment
import os
from typing import List

import openpyxl

# Import the values from the environment.py file

from environment import (
    A3_OFFSET)


class Result:
    total = 0

    def __init__(self, success: List[int], failure: List[int], timeOfExecution: int):
        self.success = success
        self.failure = failure
        # self.rsrp = rp
        self.total = sum(success) + sum(failure)
        self.environment = ""
        self.timeOfExecution = timeOfExecution

    def get_total_success(self):
        return sum(self.success)

    def get_success(self):
        return self.success

    def get_failure(self):
        return self.failure

    def get_total_failure(self):
        return sum(self.failure)

    def save_to_file(self, file_name: str, time_to_trigger: int, hysteresis: int, e_nb_location: tuple):
        # Check if the file exists
        if not os.path.exists(file_name):
            print("File does not exist, creating new file")
            # Create a new workbook and add a header row
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.cell(row=1, column=1).value = "Time of execution"
            sheet.cell(row=1, column=2).value = "Time To Trigger"
            sheet.cell(row=1, column=3).value = "Hysteresis"
            sheet.cell(row=1, column=4).value = "e_nb_location_X"
            sheet.cell(row=1, column=5).value = "e_nb_location_Y"
            sheet.cell(row=1, column=6).value = "Failed HOs"
            sheet.cell(row=1, column=7).value = "Successful HOs"
            sheet.cell(row=1, column=8).value = "Total HOs"
        else:
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        # Get the next available row on the sheet
        next_row = sheet.max_row + 1

        # Write the TTT and HYSTERESIS to the sheet
        sheet.cell(row=next_row, column=1).value = self.timeOfExecution
        sheet.cell(row=next_row, column=2).value = time_to_trigger
        sheet.cell(row=next_row, column=3).value = hysteresis

        # Write the total success and total failure to the sheet
        # sheet.cell(row=next_row, column=6).value = (self.rsrp)
        sheet.cell(row=next_row, column=4).value = (e_nb_location[0])
        sheet.cell(row=next_row, column=4).value = (e_nb_location[1])
        sheet.cell(row=next_row, column=6).value = (self.get_total_failure())
        sheet.cell(row=next_row, column=7).value = (self.get_total_success())
        sheet.cell(row=next_row, column=8).value = (self.get_total_ho())

        # Save the workbook
        workbook.save(file_name)

    def get_total_ho(self):
        return self.total
